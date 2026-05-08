"""Deterministic Excel parser using openpyxl.

Each sheet becomes a logical chunk group. Cells are extracted in stable order
(row by row, column by column).
"""

from typing import List, Optional
from .base import BaseParser
from ..models.payload import DeterministicPayload, LogicalChunk
import logging

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Deterministic Excel parser."""

    PARSER_VERSION = "0.1.0"
    PARSER_NAME = "ExcelParser"

    def parse_deterministic(
        self, file_bytes: bytes, document_id: str
    ) -> DeterministicPayload:
        """Extract deterministic payload from Excel file.

        Args:
            file_bytes: Raw XLSX file
            document_id: SHA256 of file

        Returns:
            DeterministicPayload with all sheets as logical chunks
        """
        import tempfile
        import os
        from openpyxl import load_workbook

        # Write to temp file (openpyxl can use file path)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            f.write(file_bytes)
            temp_path = f.name

        try:
            # Load workbook
            workbook = load_workbook(temp_path, data_only=True)

            chunks = []
            full_text_parts = []

            # Process each sheet in stable order
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]

                # Extract sheet as table
                sheet_text = self._extract_sheet_text(sheet, sheet_name)

                if sheet_text.strip():
                    chunk = {
                        "content": sheet_text,
                        "chunk_type": "table",
                        "source_page_start": -1,
                        "source_page_end": -1,
                        "hierarchy_level": 1,  # Top-level sheet
                        "hierarchy_path": [sheet_name],
                        "metadata": {"sheet_index": sheet_idx, "sheet_name": sheet_name},
                    }
                    chunks.append(chunk)
                    full_text_parts.append(sheet_text)

            # Combine all sheets
            full_text = "\n\n".join(full_text_parts)
            normalized_text = self._normalize_text(full_text)

            # Create logical chunks
            logical_chunks = self._create_logical_chunks(normalized_text, chunks)

            # Count statistics
            page_count = 1  # Excel doesn't have explicit pages
            word_count = len(normalized_text.split())

            # Create deterministic payload
            payload = DeterministicPayload(
                document_id=document_id,
                source_format="EXCEL",
                parser_version=self.PARSER_VERSION,
                normalization_version="1.0",
                raw_text=normalized_text,
                logical_chunks=logical_chunks,
                page_count=page_count,
                word_count=word_count,
                chunk_count=len(logical_chunks),
                encoding="utf-8",
            )

            return payload

        finally:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    def _extract_sheet_text(self, sheet, sheet_name: str) -> str:
        """Extract text from sheet in stable row/column order."""
        rows = []

        # Iterate rows
        for row in sheet.iter_rows(values_only=True):
            cells = []
            for cell_value in row:
                # Convert cell value to string
                if cell_value is None:
                    cell_text = ""
                elif isinstance(cell_value, (int, float)):
                    cell_text = str(cell_value)
                elif isinstance(cell_value, bool):
                    cell_text = "TRUE" if cell_value else "FALSE"
                else:
                    cell_text = str(cell_value)

                cells.append(cell_text)

            # Only include rows with at least one non-empty cell
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))

        # Format as table with sheet name header
        if rows:
            return f"[{sheet_name}]\n" + "\n".join(rows)
        else:
            return f"[{sheet_name}] (empty)"

    def _get_extraction_method(self) -> str:
        return "openpyxl"
