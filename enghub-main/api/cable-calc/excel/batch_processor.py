"""
ExcelBatchProcessor - пакетная обработка Excel файлов.

Процесс:
1. Открить Excel файл
2. Для каждой строки:
   a. Парсинг данных → CableInput
   b. Расчет → CalculationEngine
   c. Валидация → ValidationEngine
   d. Результат → ResultModel
3. Возврат списка ResultModel с статусами OK/WARNING/ERROR
"""
import time
import copy
from typing import List, Dict, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from engine import (
    CalculationEngine,
    ValidationEngine,
    CableInput,
    SourceParams,
)
from models import ResultModel, BatchResultSummary


class ExcelBatchProcessor:
    """Обработка Excel файлов строка за строкой."""

    # Соответствие названий колонок для парсинга
    COLUMN_MAPPING = {
        # Идентификация
        "id": ["id", "№", "номер", "line_id"],
        "name": ["name", "название", "line_name", "описание"],
        # Электрические параметры
        "section_mm2": ["сечение", "section", "section_mm2", "сечение_мм2", "s"],
        "length_m": ["длина", "length", "length_m", "длина_м", "l"],
        "power_kw": ["мощность", "power", "power_kw", "мощность_кв"],
        "phases": ["фазы", "phases", "фазность"],
        "cos_phi": ["коспи", "cos_phi", "cosφ"],
        # Параметры кабеля
        "material": ["материал", "material", "провод"],
        "insulation": ["изоляция", "insulation"],
        "method": ["способ", "method", "способ_прокладки"],
        # Условия
        "ambient_temp_c": ["температура", "temp", "ambient_temp_c", "температура_°c"],
        "cables_nearby": ["в_группе", "cables_nearby", "соседних"],
        "cable_count": ["параллельных", "cable_count", "кол-во_параллельных"],
    }

    def __init__(self):
        """Инициализация процессора."""
        self.calc_engine = CalculationEngine()
        self.validator = ValidationEngine()

    def process_batch(
        self,
        excel_path: str,
        sheet_name: Optional[str] = None,
        start_row: int = 2,  # По умолчанию первая строка - заголовок
        defaults: Optional[Dict] = None,
    ) -> Tuple[List[ResultModel], BatchResultSummary]:
        """
        Обработать Excel файл.

        Args:
            excel_path: Путь к Excel файлу
            sheet_name: Имя листа (если None, берется первый)
            start_row: Номер строки для начала парсинга (1-based, по умолчанию 2)
            defaults: Значения по умолчанию (material, insulation, method, и т.д.)

        Returns:
            (список результатов, статистика)
        """
        defaults = defaults or {}
        results = []
        t0_batch = time.time()

        try:
            # 1. Открить Excel
            wb = load_workbook(excel_path, data_only=True)

            # Выбрать лист
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # 2. Парсинг заголовков
            headers = self._parse_headers(ws)

            # 3. Обработать строки
            for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                try:
                    # Пропустить пустые строки
                    if not any(cell is not None for cell in row):
                        continue

                    # Парсинг данных строки
                    row_data = self._parse_row(row, headers)
                    if not row_data:
                        continue

                    # Создание CableInput
                    inp = self._row_to_cable_input(row_data, defaults)

                    # Валидация входных параметров
                    validation_error = CalculationEngine.validate_input(inp)
                    if validation_error:
                        result = self._error_result(
                            row_idx,
                            f"Ошибка входных данных: {validation_error}",
                            row_data,
                        )
                        results.append(result)
                        continue

                    # Расчет
                    t0 = time.time()
                    try:
                        # Выбрать режим в зависимости от входных данных
                        if inp.power_kw and inp.power_kw > 0 and not inp.section_mm2:
                            # Подбор сечения
                            calc_result = self.calc_engine.select_section(inp)
                        else:
                            # Проверка заданного сечения
                            calc_result = self.calc_engine.check_section(inp)
                    except Exception as e:
                        result = self._error_result(
                            row_idx,
                            f"Ошибка расчета: {str(e)}",
                            row_data,
                        )
                        results.append(result)
                        continue

                    calc_time_ms = (time.time() - t0) * 1000

                    # Валидация результата
                    validation = self.validator.validate(inp, calc_result)

                    # Создание единого результата
                    result = ResultModel.from_calc_result(
                        result_id=f"R{row_idx}",
                        mode="BATCH",
                        calc_input={
                            "phases": inp.phases,
                            "power_kw": inp.power_kw,
                            "length_m": inp.length_m,
                            "material": inp.material,
                            "insulation": inp.insulation,
                            "method": inp.method,
                            "section_mm2": inp.section_mm2,
                        },
                        calc_result={
                            "section_mm2": calc_result.section_mm2,
                            "i_allowable_a": calc_result.i_allowable_a,
                            "i_calc_a": calc_result.i_calc_a,
                            "delta_u_pct": calc_result.delta_u_pct,
                            "cb_rating_a": calc_result.cb_rating_a,
                            "status": calc_result.status,
                        },
                        validation=validation,
                        original_data=row_data,
                        calc_time_ms=calc_time_ms,
                    )

                    # Добавить сведения о расчете
                    result.notes = (
                        "Подбор сечения"
                        if not inp.section_mm2
                        else "Проверка сечения"
                    )

                    results.append(result)

                except Exception as e:
                    result = self._error_result(
                        row_idx,
                        f"Неожиданная ошибка: {str(e)}",
                        {},
                    )
                    results.append(result)

        except Exception as e:
            # Общая ошибка при откритии файла
            return (
                [self._error_result(0, f"Ошибка открытия файла: {str(e)}")],
                BatchResultSummary(),
            )

        # 4. Подсчет статистики
        t1_batch = time.time()
        summary = BatchResultSummary(
            total_processed=len(results),
            ok_count=sum(1 for r in results if r.status == "OK"),
            warning_count=sum(1 for r in results if r.status == "WARNING"),
            error_count=sum(1 for r in results if r.status == "ERROR"),
            total_time_ms=(t1_batch - t0_batch) * 1000,
        )

        return results, summary

    def _parse_headers(self, ws) -> Dict[str, int]:
        """Парсинг заголовков Excel (первая строка)."""
        headers = {}
        for col_idx, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            if cell[0].value:
                header_text = str(cell[0].value).lower().strip()
                headers[header_text] = col_idx
        return headers

    def _parse_row(self, row: Tuple, headers: Dict[str, int]) -> Optional[Dict]:
        """Парсинг строки Excel в словарь данных."""
        row_data = {}

        # Попробовать найти каждый параметр в заголовках
        for param_name, column_aliases in self.COLUMN_MAPPING.items():
            for alias in column_aliases:
                if alias in headers:
                    col_idx = headers[alias] - 1  # Конвертировать в 0-based индекс
                    if col_idx < len(row):
                        value = row[col_idx]
                        if value is not None:
                            row_data[param_name] = value
                    break

        return row_data if row_data else None

    def _row_to_cable_input(self, row_data: Dict, defaults: Dict) -> CableInput:
        """Преобразование строки Excel в CableInput."""
        src = SourceParams()

        # Парсинг фаз
        phases = self._parse_int(row_data.get("phases", defaults.get("phases", 3)))
        if phases not in (1, 3):
            phases = 3

        # Парсинг мощности
        power_kw = self._parse_float(row_data.get("power_kw"))

        # Парсинг сечения
        section_mm2 = self._parse_float(row_data.get("section_mm2"))

        # Парсинг длины
        length_m = self._parse_float(row_data.get("length_m", defaults.get("length_m", 50.0)))
        if length_m <= 0:
            length_m = 50.0

        # Материал (Cu/Al)
        material = str(row_data.get("material", defaults.get("material", "Cu"))).upper()
        if material not in ("CU", "AL"):
            material = "Cu"

        # Изоляция (PVC/XLPE)
        insulation = str(row_data.get("insulation", defaults.get("insulation", "PVC"))).upper()
        if insulation not in ("PVC", "XLPE"):
            insulation = "PVC"

        # Способ прокладки
        method = str(row_data.get("method", defaults.get("method", "C"))).upper()
        if method not in ("A1", "A2", "B1", "B2", "C", "D1", "D2", "E", "F", "G"):
            method = "C"

        # Кол-во в группе
        cables_nearby = self._parse_int(row_data.get("cables_nearby", 1))
        if cables_nearby < 1:
            cables_nearby = 1

        # Параллельные кабели
        cable_count = self._parse_int(row_data.get("cable_count", 1))
        if cable_count < 1:
            cable_count = 1

        # Температура
        ambient_temp_c = self._parse_float(row_data.get("ambient_temp_c", 30.0))
        if ambient_temp_c < -40 or ambient_temp_c > 70:
            ambient_temp_c = 30.0

        # cos φ
        cos_phi = self._parse_float(row_data.get("cos_phi", 0.85))
        if cos_phi <= 0 or cos_phi > 1:
            cos_phi = 0.85

        return CableInput(
            line_id=str(row_data.get("id", "")),
            line_name=str(row_data.get("name", "")),
            phases=phases,
            power_kw=power_kw,
            cos_phi=cos_phi,
            length_m=length_m,
            material=material,
            insulation=insulation,
            method=method,
            cables_nearby=cables_nearby,
            cable_count=cable_count,
            section_mm2=section_mm2,
            ambient_temp_c=ambient_temp_c,
            source=src,
        )

    @staticmethod
    def _parse_float(value) -> Optional[float]:
        """Безопасное преобразование в float."""
        if value is None:
            return None
        try:
            return float(str(value).replace(",", "."))
        except (ValueError, AttributeError):
            return None

    @staticmethod
    def _parse_int(value) -> int:
        """Безопасное преобразование в int."""
        if value is None:
            return 0
        try:
            return int(float(str(value).replace(",", ".")))
        except (ValueError, AttributeError):
            return 0

    @staticmethod
    def _error_result(row_idx: int, error_msg: str, original_data: Optional[Dict] = None) -> ResultModel:
        """Создать результат с ошибкой."""
        return ResultModel(
            id=f"R{row_idx}",
            mode="BATCH",
            timestamp=__import__("datetime").datetime.now().isoformat(),
            input={},
            calculated={},
            original=original_data or {},
            status="ERROR",
            issues=[error_msg],
            warnings=[],
            calculation_time_ms=0.0,
        )
