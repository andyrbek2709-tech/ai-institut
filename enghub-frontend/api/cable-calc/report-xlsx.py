"""
Vercel Python serverless function — генерация Excel-отчёта по результатам массовой проверки журнала.

POST /api/cable-calc/report-xlsx
JSON: {
  "source_file": "журнал_кабелей.xlsx",
  "defaults": {"material": "Cu", "insulation": "PVC", "method": "C", "ambient_temp_c": 30},
  "lines": [
    {row_num, cable_id, cable_name, from_point, to_point, cable_mark, section_str,
     section_mm2, zero_section_mm2, phases, length_m, voltage_kv,
     i_allowable_a, status, note}, ...
  ]
}

Ответ: .xlsx файл с теми же колонками + результат проверки. Раскраска строк по статусу.
"""
import io
import json
import time
import traceback
from http.server import BaseHTTPRequestHandler

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception as e:
    Workbook = None
    _IMPORT_ERR = str(e)


HEADERS = [
    ("№",                "row_num",          6),
    ("Обозначение",      "cable_id",         18),
    ("Наименование",     "cable_name",       28),
    ("Начало",           "from_point",       22),
    ("Конец",            "to_point",         22),
    ("Марка кабеля",     "cable_mark",       18),
    ("Сечение",          "section_str",      14),
    ("S, мм²",           "section_mm2",      9),
    ("S₀, мм²",          "zero_section_mm2", 9),
    ("Длина, м",         "length_m",         10),
    ("U, кВ",            "voltage_kv",       8),
    ("I_доп, А",         "i_allowable_a",    11),
    ("Статус",           "status",           10),
    ("Комментарий",      "note",             40),
]


def _build_xlsx(payload: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал"

    source_file = payload.get("source_file", "")
    defaults = payload.get("defaults", {}) or {}
    lines = payload.get("lines", []) or []

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    sub_font = Font(italic=True, color="555555", size=10)

    fill_ok = PatternFill("solid", fgColor="DCFCE7")     # зелёный
    fill_warn = PatternFill("solid", fgColor="FEF3C7")   # жёлтый
    fill_fail = PatternFill("solid", fgColor="FEE2E2")   # красный

    # Заголовок отчёта
    last_col_letter = get_column_letter(len(HEADERS))
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "Проверка кабельного журнала — EngHub"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    sub = (
        f"Файл: {source_file}    |    "
        f"Материал: {defaults.get('material', 'Cu')}    "
        f"Изоляция: {defaults.get('insulation', 'PVC')}    "
        f"Метод: {defaults.get('method', 'C')}    "
        f"t_среды: {defaults.get('ambient_temp_c', 30)}°C    "
        f"Дата: {time.strftime('%d.%m.%Y %H:%M')}"
    )
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = sub
    ws["A2"].font = sub_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # Шапка таблицы
    header_row = 4
    for col_idx, (label, _key, width) in enumerate(HEADERS, 1):
        c = ws.cell(row=header_row, column=col_idx, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 30

    # Данные
    ok_cnt = warn_cnt = fail_cnt = 0
    for r_offset, line in enumerate(lines, 1):
        row = header_row + r_offset
        status = (line.get("status") or "").upper()
        if status == "OK":
            fill = fill_ok
            ok_cnt += 1
        elif status in ("FAIL", "ERROR"):
            fill = fill_fail
            fail_cnt += 1
        else:
            fill = fill_warn
            warn_cnt += 1

        for col_idx, (_label, key, _w) in enumerate(HEADERS, 1):
            v = line.get(key, "")
            if isinstance(v, float) and v == int(v):
                v = int(v)
            c = ws.cell(row=row, column=col_idx, value=v)
            c.fill = fill
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=(key in ("note", "cable_name", "from_point", "to_point")))
            if key == "status":
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")

    # Итоговая строка
    summary_row = header_row + len(lines) + 2
    ws.cell(row=summary_row, column=1, value="Итого:").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"строк {len(lines)}")
    ws.cell(row=summary_row, column=3, value=f"OK: {ok_cnt}").fill = fill_ok
    ws.cell(row=summary_row, column=4, value=f"WARN: {warn_cnt}").fill = fill_warn
    ws.cell(row=summary_row, column=5, value=f"FAIL: {fail_cnt}").fill = fill_fail

    # Зафиксировать шапку
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class handler(BaseHTTPRequestHandler):
    def _send_json(self, status: int, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self._send_json(204, {})

    def do_GET(self):
        self._send_json(200, {
            "ok": True,
            "service": "cable-calc/report-xlsx",
            "version": "1.0",
            "format": "xlsx",
        })

    def do_POST(self):
        try:
            if Workbook is None:
                return self._send_json(500, {"error": f"openpyxl not available: {_IMPORT_ERR}"})
            length = int(self.headers.get("Content-Length", "0"))
            raw = self.rfile.read(length) if length > 0 else b"{}"
            payload = json.loads(raw.decode("utf-8") or "{}")
            data = _build_xlsx(payload)
            ts = time.strftime("%Y%m%d_%H%M%S")
            filename = f"cable_journal_check_{ts}.xlsx"

            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header("Content-Length", str(len(data)))
            self.send_header(
                "Content-Disposition",
                f'attachment; filename="{filename}"',
            )
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Expose-Headers", "Content-Disposition")
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            self._send_json(500, {
                "error": str(e),
                "type": type(e).__name__,
                "traceback": traceback.format_exc().splitlines()[-8:],
            })
